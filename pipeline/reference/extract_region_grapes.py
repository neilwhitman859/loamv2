"""
Extract grape varieties by region from Anderson & Aryal dataset (2023).

Source: University of Adelaide Wine Economics Research Centre
"Database of Regional, National and Global Winegrape Bearing Areas by Variety, 2000 to 2023"
https://economics.adelaide.edu.au/wine-economics/databases

Outputs JSON with top grapes per region and per country for seeding
region_grapes and country_grapes tables.

Usage:
    python -m pipeline.reference.extract_region_grapes --regional <file.xlsx> [--national <file.xlsx>]
"""

import argparse
import json
import sys
from pathlib import Path

# Allow running as module from repo root
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]

# Thresholds
MIN_HECTARES_REGION = 50
MIN_SHARE_REGION = 0.03  # 3%
MAX_GRAPES_PER_REGION = 15
MIN_HECTARES_COUNTRY = 500
MIN_SHARE_COUNTRY = 0.02  # 2%
MAX_GRAPES_PER_COUNTRY = 20


def read_xlsx_sheets(path: str) -> dict[str, list[dict]]:
    """Read all sheets from an XLSX file, returning {sheet_name: [row_dicts]}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        if name in ("Title page", "All countries"):
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        sheet_data = []
        for row in rows[1:]:
            row_dict = dict(zip(headers, row))
            sheet_data.append(row_dict)
        sheets[name] = sheet_data
    wb.close()
    return sheets


def main():
    parser = argparse.ArgumentParser(
        description="Extract grape-by-region data from Anderson & Aryal Excel"
    )
    parser.add_argument("--regional", required=True, help="Path to regional XLSX file")
    parser.add_argument("--national", help="Path to national XLSX file (optional)")
    parser.add_argument("--output", default=str(PROJECT_ROOT / "data" / "anderson_aryal_grapes.json"),
                        help="Output JSON file")
    args = parser.parse_args()

    # Read regional data
    print(f"Reading regional file: {args.regional}")
    sheets = read_xlsx_sheets(args.regional)

    region_data: dict[str, dict[str, dict[str, float]]] = {}  # country -> region -> grape -> ha
    country_data: dict[str, dict[str, float]] = {}  # country -> grape -> ha

    for sheet_name, rows in sheets.items():
        country = sheet_name
        if country not in country_data:
            country_data[country] = {}
        if country not in region_data:
            region_data[country] = {}

        for row in rows:
            region = row.get("region", "Unknown") or "Unknown"
            grape = row.get("prime")
            area = 0.0
            try:
                area = float(row.get("area", 0) or 0)
            except (ValueError, TypeError):
                pass

            if not grape or area <= 0:
                continue

            # Aggregate by region
            if region not in region_data[country]:
                region_data[country][region] = {}
            region_data[country][region][grape] = region_data[country][region].get(grape, 0) + area

            # Aggregate by country
            country_data[country][grape] = country_data[country].get(grape, 0) + area

    # Read national data if provided
    if args.national:
        print(f"Reading national file: {args.national}")
        nat_sheets = read_xlsx_sheets(args.national)
        print(f"National file sheets: {list(nat_sheets.keys())[:5]}")
        # Show structure of first data sheet
        first_key = list(nat_sheets.keys())[0] if nat_sheets else None
        if first_key:
            first_rows = nat_sheets[first_key][:5]
            print("National first rows:")
            for i, r in enumerate(first_rows):
                vals = list(r.values())[:10]
                print(f"  {i}: {vals}")

    # Process into output
    output = {
        "source": (
            "Anderson, Nelgen & Puga (2023). Database of Regional, National and Global "
            "Winegrape Bearing Areas by Variety, 2000 to 2023. "
            "University of Adelaide Wine Economics Research Centre."
        ),
        "url": "https://economics.adelaide.edu.au/wine-economics/databases",
        "countries": {},
        "regions": {},
    }

    # Country-level: top grapes by hectares
    for country, grapes in country_data.items():
        sorted_grapes = sorted(grapes.items(), key=lambda x: -x[1])
        total_ha = sum(ha for _, ha in sorted_grapes)
        if total_ha <= 0:
            continue

        top_grapes = [
            {
                "grape": grape,
                "hectares": round(ha),
                "share": f"{round((ha / total_ha) * 1000) / 10}%",
            }
            for grape, ha in sorted_grapes
            if ha >= MIN_HECTARES_COUNTRY and (ha / total_ha) >= MIN_SHARE_COUNTRY
        ][:MAX_GRAPES_PER_COUNTRY]

        output["countries"][country] = {
            "total_hectares": round(total_ha),
            "grapes": top_grapes,
        }

    # Region-level: top grapes by hectares
    for country, regions in region_data.items():
        for region, grapes in regions.items():
            sorted_grapes = sorted(grapes.items(), key=lambda x: -x[1])
            total_ha = sum(ha for _, ha in sorted_grapes)
            if total_ha <= 0:
                continue

            top_grapes = [
                {
                    "grape": grape,
                    "hectares": round(ha),
                    "share": f"{round((ha / total_ha) * 1000) / 10}%",
                }
                for grape, ha in sorted_grapes
                if ha >= MIN_HECTARES_REGION and (ha / total_ha) >= MIN_SHARE_REGION
            ][:MAX_GRAPES_PER_REGION]

            if top_grapes:
                key = f"{country}|{region}"
                output["regions"][key] = {
                    "country": country,
                    "region": region,
                    "total_hectares": round(total_ha),
                    "grapes": top_grapes,
                }

    # Write output
    out_path = Path(args.output)
    out_path.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"\n=== SUMMARY ===")
    print(f"Countries: {len(output['countries'])}")
    print(f"Regions: {len(output['regions'])}")

    # Print country summaries
    print(f"\n=== COUNTRY GRAPES (top 5 each) ===")
    for country, data in sorted(output["countries"].items(), key=lambda x: -x[1]["total_hectares"]):
        top_names = ", ".join(f"{g['grape']} ({g['share']})" for g in data["grapes"][:5])
        print(f"{country} [{data['total_hectares']} ha]: {top_names}")

    print(f"\n=== REGIONS WITH DATA ===")
    for key in sorted(output["regions"].keys()):
        data = output["regions"][key]
        top_names = ", ".join(g["grape"] for g in data["grapes"][:3])
        print(f"  {data['country']} > {data['region']}: {len(data['grapes'])} grapes ({top_names}...)")


if __name__ == "__main__":
    main()
