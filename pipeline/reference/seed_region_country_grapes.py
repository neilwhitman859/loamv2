"""
Seed region_grapes and country_grapes from Anderson & Aryal dataset.

Source: Anderson, Nelgen & Puga (2023). "Database of Regional, National and
Global Winegrape Bearing Areas by Variety, 2000 to 2023."
University of Adelaide Wine Economics Research Centre.
https://economics.adelaide.edu.au/wine-economics/databases

Strategy:
1. Parse the regional Excel file for grape plantings by region
2. Match Anderson regions to our Loam regions (fuzzy matching)
3. Match Anderson grape names to our VIVC-based grapes (synonym matching)
4. Generate SQL inserts for region_grapes and country_grapes
5. Clear existing data first (we're rebuilding from authoritative source)

Usage:
    python -m pipeline.reference.seed_region_country_grapes --regional <file.xlsx> [--dry-run]
"""

import argparse
import json
import sys
from pathlib import Path

# Allow running as module from repo root
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import openpyxl

from pipeline.lib.db import get_supabase, fetch_all
from pipeline.lib.normalize import normalize

PROJECT_ROOT = Path(__file__).resolve().parents[2]

# Thresholds
MIN_HA_REGION = 50
MIN_SHARE_REGION = 0.02  # 2% of region total
MAX_GRAPES_REGION = 15
MIN_HA_COUNTRY = 200
MIN_SHARE_COUNTRY = 0.015  # 1.5% of country total
MAX_GRAPES_COUNTRY = 25

# Anderson country name -> our ISO code mapping
COUNTRY_MAP = {
    "Albania": "AL", "Algeria": None, "Argentina": "AR", "Armenia": "AM",
    "Australia": "AU", "Austria": "AT", "Belgium": "BE", "Brazil": "BR",
    "Bulgaria": "BG", "Cambodia": None, "Canada": "CA", "Chile": "CL",
    "China": "CN", "Croatia": "HR", "Cyprus": "CY", "Czechia": "CZ",
    "Ethiopia": None, "France": "FR", "Georgia": "GE", "Germany": "DE",
    "Greece": "GR", "Hungary": "HU", "India": "IN", "Israel": "IL",
    "Italy": "IT", "Japan": "JP", "Kazakhstan": None, "Korea, Rep.": None,
    "Lebanon": "LB", "Lithuania": None, "Luxembourg": "LU", "Mexico": "MX",
    "Moldova": "MD", "Morocco": "MA", "Myanmar": "MM", "New Zealand": "NZ",
    "North Macedonia": "MK", "Norway": None, "Peru": "PE", "Poland": "PL",
    "Portugal": "PT", "Romania": "RO", "Russia": "RU", "Serbia": "RS",
    "Slovakia": "SK", "Slovenia": "SI", "South Africa": "ZA", "Spain": "ES",
    "Sweden": "SE", "Switzerland": "CH", "Taiwan": None, "Thailand": "TH",
    "Tunisia": "TN", "Turkiye": "TR", "Turkmenistan": None, "Ukraine": "UA",
    "United Kingdom": "GB", "United States": "US", "Uruguay": "UY",
}

# Anderson region name -> our region name mapping (where they differ)
REGION_MAP = {
    # France -- Anderson uses admin regions, we use wine regions
    "FR|Alsace": "Alsace",
    "FR|Aquitaine": "Bordeaux",
    "FR|Bourgogne": "Burgundy",
    "FR|Champagne": "Champagne",
    "FR|Centre": "Loire Valley",
    "FR|Pays De Loire": "Loire Valley",
    "FR|Corse": "Corsica",
    "FR|Languedoc-Roussillon": "Southern France",
    "FR|Provence -C. D'Azur": "Southern France",
    "FR|Rhone-Alpes": "Rh\u00f4ne Valley",
    "FR|Midi-Pyrenees": "Southwest France",
    "FR|Mourvedre N": "Southwest France",
    "FR|Poitou-Charentes": "Cognac",
    # Italy -- Anderson uses admin regions
    "IT|Lombardia": "Lombardy",
    "IT|Piemonte": "Piemonte",
    "IT|Toscana": "Tuscany",
    "IT|Sicilia": "Sicily",
    "IT|Sardegna": "Sardinia",
    "IT|Provincia Autonoma di Bolzano/Bozen": "Trentino-Alto Adige",
    "IT|Provincia Autonoma di Trento": "Trentino-Alto Adige",
    "IT|Puglia": "Puglia",
    "IT|Veneto": "Veneto",
    "IT|Emilia-Romagna": "Emilia-Romagna",
    "IT|Friuli-Venezia Giulia": "Friuli-Venezia Giulia",
    "IT|Campania": "Campania",
    "IT|Abruzzo": "Abruzzo",
    "IT|Lazio": "Lazio",
    "IT|Calabria": "Calabria",
    "IT|Basilicata": "Basilicata",
    "IT|Liguria": "Liguria",
    "IT|Umbria": "Umbria",
    "IT|Marche": "Marche",
    "IT|Molise": "Molise",
    # Germany
    "DE|Trier": "Mosel",
    "DE|Koblenz": "Mosel",
    "DE|Rheinhessen-Pfalz": "Rheinhessen",
    "DE|Freiburg": "Baden",
    "DE|Karlsruhe": "Baden",
    "DE|Stuttgart": "W\u00fcrttemberg",
    "DE|T\u00fcbingen": "W\u00fcrttemberg",
    "DE|Darmstadt": "Rheingau",
    "DE|Mittelfranken": "Franken",
    "DE|Unterfranken": "Franken",
    "DE|Sachsen-Anhalt": "Saale-Unstrut",
    # Spain
    "ES|La Rioja": "The Upper Ebro",
    "ES|Navarra": "The Upper Ebro",
    "ES|Galicia": "The North West",
    "ES|Catalu\u00f1a": "Catalunya",
    "ES|Castilla Y Le\u00f3n": "Castilla y Le\u00f3n",
    "ES|Castilla La\nMancha": "Castilla-La Mancha",
    "ES|Arag\u00f3n": "Arag\u00f3n",
    "ES|Andaluc\u00eda": "Andaluc\u00eda",
    "ES|Valencia": "The Levante",
    "ES|Murcia": "The Levante",
    "ES|Madrid": "Madrid",
    "ES|Extremadura": "Extremadura",
    "ES|Pa\u00eds Vasco": "The Upper Ebro",
    "ES|Baleares": "Balearic Islands",
    "ES|Canarias": "Canary Islands",
    "ES|Castilla La Mancha": "Castilla-La Mancha",
    "ES|Asturias": None,
    "ES|Cantabria": None,
    # Chile
    "CL|Valparaiso": "Aconcagua Region",
    "CL|Metropolitana": "Aconcagua Region",
    "CL|L.B.O'Higgins": "Central Valley Region",
    "CL|Maule": "Central Valley Region",
    "CL|Coquimbo": "Coquimbo Region",
    "CL|Bio Bio": "Southern Region",
    "CL|\u00d1uble": "Southern Region",
    "CL|Araucania": "Southern Region",
    "CL|Atacama": "Atacama",
    # Greece
    "GR|Kentriki Makedonia": "Macedonia",
    "GR|Dytiki Makedonia": "Macedonia",
    "GR|Anatoliki Makedonia, Thraki": "Macedonia",
    "GR|Peloponnisos": "Peloponnese",
    "GR|Dytiki Ell\u00e1da": "Peloponnese",
    "GR|Attiki": "Central Greece",
    "GR|Sterea Ell\u00e1da": "Central Greece",
    "GR|Thessalia": "Central Greece",
    "GR|Kriti": "Crete",
    "GR|Notio Aigaio": "Aegean Islands",
    "GR|Voreio Aigaio": "Aegean Islands",
    "GR|Ionia Nisia": "Ionian Islands",
    "GR|Ipeiros": "Epirus",
    # Hungary
    "HU|\u00c9szak-Magyarorsz\u00e1g": None,
    "HU|D\u00e9l-Dun\u00e1nt\u00fal": None,
    "HU|Nyugat-Dun\u00e1nt\u00fal": None,
    "HU|D\u00e9l-Alf\u00f6ld": None,
    "HU|\u00c9szak-Alf\u00f6ld": None,
    "HU|K\u00f6z\u00e9p-Dun\u00e1nt\u00fal": "Lake Balaton",
    "HU|Pest": None,
    "HU|Budapest": None,
    # Slovenia -- skip all (no L2 regions)
    "SI|Gori\u0161ka Brda": None,
    "SI|Vipavska Dolina": None,
    "SI|Slovenska Istra": None,
    "SI|\u0160tajerska Slovenija": None,
    "SI|Kras": None,
    "SI|Dolenjska": None,
    "SI|Bizeljsko-Sremi\u010d": None,
    "SI|Bela Krajina": None,
    "SI|Prekmurje": None,
    "SI|Ni Okoli\u0161": None,
    # Portugal
    "PT|Douro": "Douro",
    "PT|Alentejo": "Alentejo",
    "PT|Minho": "Vinho Verde",
    "PT|Lisboa": "Lisboa",
    "PT|Tejo": "Tejo",
    "PT|Beira Atl\u00e2ntico": "Bairrada",
    "PT|Terras Do D\u00e3o": "D\u00e3o",
    "PT|Terras De Cister": "D\u00e3o",
    "PT|Terras Da Beira": "Beira Interior",
    "PT|Tr\u00e1s-Os-Montes": "Tr\u00e1s-os-Montes",
    "PT|Algarve": "Algarve",
    "PT|Pen\u00ednsula De Set\u00fabal": "Set\u00fabal",
    # Australia
    "AU|SA": "South Australia",
    "AU|Vic": "Victoria",
    "AU|NSW": "New South Wales",
    "AU|WA": "Western Australia",
    "AU|Tas": "Tasmania",
    "AU|Qld": "Queensland",
    # NZ
    "NZ|Marlborough": "South Island",
    "NZ|Central Otago": "South Island",
    "NZ|North Canterbury": "South Island",
    "NZ|Nelson": "South Island",
    "NZ|Hawkes Bay": "North Island",
    "NZ|Gisborne": "North Island",
    "NZ|Auckland": "North Island",
    "NZ|Wairarapa": "North Island",
    # South Africa -- handle newlines in names
    "ZA|Stellen-\nBosch": "Coastal Region", "ZA|Stellen- Bosch": "Coastal Region",
    "ZA|Paarl": "Coastal Region",
    "ZA|Swart-\nLand": "Coastal Region", "ZA|Swart- Land": "Coastal Region",
    "ZA|Cape\nTown": "Coastal Region", "ZA|Cape Town": "Coastal Region",
    "ZA|Cape South Coast": "Cape South Coast",
    "ZA|Robert-\nSon": "Breede River Valley", "ZA|Robert- Son": "Breede River Valley",
    "ZA|Worces-\nTer": "Breede River Valley", "ZA|Worces- Ter": "Breede River Valley",
    "ZA|Breede-\nKloof": "Breede River Valley", "ZA|Breede- Kloof": "Breede River Valley",
    "ZA|Klein\nKaroo": "Western Cape", "ZA|Klein Karoo": "Western Cape",
    "ZA|Olifants\nRiver": "Western Cape", "ZA|Olifants River": "Western Cape",
    "ZA|Northern\nCape": "Western Cape", "ZA|Northern Cape": "Western Cape",
    # US
    "US|California": "California",
    "US|Washington": "Washington",
    "US|Oregon": "Oregon",
    "US|New York": "New York",
    "US|Texas": "Texas",
    # Switzerland
    "CH|VD": "Vaud",
    "CH|VS": "Valais",
    "CH|GE": "Geneva",
    "CH|TI": "Ticino",
    "CH|ZH": "Z\u00fcrich",
    "CH|NE": "Neuch\u00e2tel",
    "CH|SH": "Schaffhausen",
    "CH|TG": "Thurgau",
    "CH|AG": "Aargau",
    "CH|SG": "St. Gallen",
    "CH|BE Lac de Bienne": "Bern",
    "CH|GR \u00fcbriges Gebiet": "Graub\u00fcnden",
    "CH|BL": None,
    # Austria
    "AT|Nieder\u00f6sterreich": "Nieder\u00f6sterreich",
    "AT|Burgenland": "Burgenland",
    "AT|Steiermark": "Steiermark",
    "AT|Wien": "Wien",
    # Georgia
    "GE|Kakheti": "Kakheti",
    "GE|Racha And Lechkhumi": "Racha-Lechkhumi",
    # Canada
    "CA|British Columbia": "British Columbia",
    # Argentina
    "AR|Mendoza": "Mendoza",
    "AR|Salta": "Salta",
    "AR|San Juan": "San Juan",
    "AR|La Rioja": "La Rioja",
    "AR|Rio Negro": "Patagonia",
    "AR|Neuquen": "Patagonia",
    "AR|Catamarca": None,
    "AR|Cordoba": "C\u00f3rdoba",
    "AR|Tucuman": None,
    "AR|La Pampa": None,
    "AR|Buenos Aires": "Buenos Aires",
    "AR|Jujuy": "Jujuy",
    "AR|Entre Rios": "Entre R\u00edos",
    "AR|San Luis": None,
    "AR|Chubut": "Patagonia",
    "AR|Misiones": None,
    "AR|Santa Fe": None,
    "AR|S Del Estero": None,
    # US -- skip non-vinifera
    "US|Michigan": None,
    # Japan
    "JP|Yamanashi": None,
    "JP|Hokkaido": None,
    "JP|Nagano": None,
    # Romania/Bulgaria/Slovakia/Russia/Czechia/Poland -- skip admin regions
    "SK|Bratislavsk\u00fd kraj": None,
    "SK|Z\u00e1padn\u00e9 Slovensko": None,
    "SK|Stredn\u00e9 Slovensko": None,
    "SK|V\u00fdchodn\u00e9 Slovensko": None,
    "RU|Crimea": None,
    "RU|Krasnodar Krai": None,
    "RU|Other regions": None,
    "RU|Rostov Oblast": None,
    "CZ|Morava": None,
    "CZ|\u010cechy": None,
}


def read_xlsx_sheets(path: str) -> dict[str, list[dict]]:
    """Read all sheets from an XLSX file, returning {sheet_name: [row_dicts]}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        if name in ("Title page", "All countries"):
            continue
        ws = wb[name]
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            continue
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows_raw[0])]
        sheet_data = []
        for row in rows_raw[1:]:
            row_dict = dict(zip(headers, row))
            sheet_data.append(row_dict)
        sheets[name] = sheet_data
    wb.close()
    return sheets


def main():
    parser = argparse.ArgumentParser(
        description="Seed region_grapes and country_grapes from Anderson & Aryal"
    )
    parser.add_argument("--regional", required=True, help="Path to regional XLSX file")
    parser.add_argument("--dry-run", action="store_true", help="Preview without DB writes")
    args = parser.parse_args()

    sb = get_supabase()

    # 1. Load our DB data
    print("Loading DB data...")

    countries = fetch_all("countries", "id,name,iso_code")
    print(f"  Loaded {len(countries)} countries")
    country_by_iso = {c["iso_code"]: c for c in countries}

    # Get regions (non-catch-all)
    all_regions = fetch_all("regions", "id,name,country_id,is_catch_all", {"is_catch_all": False})
    print(f"  {len(countries)} countries, {len(all_regions)} regions")

    # Build region lookup: { country_id: { name_lower: region } }
    region_lookup: dict[str, dict[str, dict]] = {}
    for r in all_regions:
        cid = r["country_id"]
        if cid not in region_lookup:
            region_lookup[cid] = {}
        region_lookup[cid][r["name"].lower()] = r

    # Get grapes with display_name and synonyms
    all_grapes = fetch_all("grapes", "id,name,display_name")
    print(f"  {len(all_grapes)} grapes")

    all_synonyms = fetch_all("grape_synonyms", "grape_id,synonym")
    print(f"  {len(all_synonyms)} synonyms")

    # Build grape lookup by various name forms
    grape_lookup: dict[str, str] = {}  # lowered name -> grape_id
    for g in all_grapes:
        grape_lookup[g["name"].lower()] = g["id"]
        if g.get("display_name"):
            grape_lookup[g["display_name"].lower()] = g["id"]
    for s in all_synonyms:
        key = s["synonym"].lower()
        if key not in grape_lookup:
            grape_lookup[key] = s["grape_id"]

    def lookup_grape(vivc_name: str) -> str | None:
        g = next((g for g in all_grapes if g["name"] == vivc_name), None)
        return g["id"] if g else None

    # Manual grape name overrides for Anderson names that don't match
    GRAPE_OVERRIDES = {
        "côt": grape_lookup.get("malbec") or grape_lookup.get("cot"),
        "douce noire": grape_lookup.get("corbeau"),
        "tribidrag": grape_lookup.get("primitivo"),
        "prosecco": grape_lookup.get("glera"),
        "syrah": grape_lookup.get("syrah"),
        "graševina": grape_lookup.get("graševina") or grape_lookup.get("welschriesling") or lookup_grape("WELSCHRIESLING"),
        "mammolo": lookup_grape("MAMMOLO"),
        "nero d'avola": grape_lookup.get("nero d'avola") or grape_lookup.get("calabrese"),
        "listán prieto": grape_lookup.get("listan prieto") or lookup_grape("LISTAN PRIETO"),
        "listán de huelva": grape_lookup.get("listán de huelva") or lookup_grape("PALOMINO FINO"),
        "listán negro": lookup_grape("LISTAN NEGRO"),
        "muscat blanc à petits grains": grape_lookup.get("muscat blanc à petits grains") or lookup_grape("MUSCAT A PETITS GRAINS BLANCS"),
        "palomino fino": lookup_grape("PALOMINO FINO"),
        "pedro ximénez": lookup_grape("PEDRO XIMENEZ"),
        "schiava grossa": lookup_grape("SCHIAVA GROSSA"),
        "mazuelo": grape_lookup.get("carignan noir"),
        "alicante henri bouschet": lookup_grape("ALICANTE HENRI BOUSCHET"),
        "catarratto bianco": grape_lookup.get("catarratto bianco comune"),
        "monastrell": grape_lookup.get("mourvèdre") or grape_lookup.get("monastrell"),
        "macabeo": lookup_grape("MACABEO"),
        "airén": lookup_grape("AIREN"),
        "bobal": lookup_grape("BOBAL"),
        "hondarribi zuri": grape_lookup.get("courbu blanc"),
        "castelão": lookup_grape("CASTELAO"),
        "alvarinho": grape_lookup.get("albariño") or grape_lookup.get("alvarinho"),
        "arinto de bucelas": lookup_grape("ARINTO"),
        "loureiro": lookup_grape("LOUREIRO"),
        "rufete": lookup_grape("RUFETE"),
        "malvasia fina": lookup_grape("MALVASIA FINA"),
        "mencía": grape_lookup.get("mencía") or lookup_grape("MENCIA"),
        "trebbiano romagnolo": lookup_grape("TREBBIANO ROMAGNOLO"),
        "lambrusco salamino": lookup_grape("LAMBRUSCO SALAMINO"),
        "falanghina flegrea": lookup_grape("FALANGHINA"),
        "teroldego": lookup_grape("TEROLDEGO"),
        "bonarda piemontese": lookup_grape("BONARDA PIEMONTESE"),
        "grechetto di orvieto": lookup_grape("GRECHETTO"),
        "gaglioppo": lookup_grape("GAGLIOPPO"),
        "magliocco canino": lookup_grape("MAGLIOCCO CANINO"),
        "trebbiano d'abruzzo": lookup_grape("TREBBIANO ABRUZZESE"),
        "malvasia bianca di candia": lookup_grape("MALVASIA BIANCA DI CANDIA"),
        "muscat of hamburg": lookup_grape("MUSCAT HAMBURG"),
        "fetească regală": lookup_grape("FETEASCA REGALA"),
        "fetească albă": lookup_grape("FETEASCA ALBA"),
        "vranac": lookup_grape("VRANAC"),
        "dimyat": lookup_grape("DIMIAT"),
        "pamid": lookup_grape("PAMID"),
        "muscat ottonel": lookup_grape("MUSCAT OTTONEL"),
        "isabella": lookup_grape("ISABELLA"),
        "siroka melniska": lookup_grape("SIROKA MELNISHKA LOZA"),
        "xynisteri": lookup_grape("XYNISTERI"),
        "sultaniye": lookup_grape("SULTANIYE"),
        "plavac mali crni": lookup_grape("PLAVAC MALI"),
        "malvazija istarska": lookup_grape("MALVAZIJA ISTARSKA"),
        "kotsifali": lookup_grape("KOTSIFALI"),
        "mavrodafni": lookup_grape("MAVRODAPHNI"),
        "mandilaria": lookup_grape("MANDILARIA"),
        "savatiano": lookup_grape("SAVATIANO"),
        "liatiko": lookup_grape("LIATIKO"),
        "dornfelder": lookup_grape("DORNFELDER"),
        "bacchus": lookup_grape("BACCHUS WEISS"),
        "xarello": lookup_grape("XAREL-LO"),
        "parellada": lookup_grape("PARELLADA"),
        "cayetana blanca": lookup_grape("CAYETANA BLANCA"),
        "manto negro": lookup_grape("MANTO NEGRO"),
        "verdejo": grape_lookup.get("verdejo") or lookup_grape("VERDEJO BLANCO"),
        "cserszegi fűszeres": lookup_grape("CSERSZEGI FUESZERES"),
        "bianca": lookup_grape("BIANCA"),
        "ribolla gialla": lookup_grape("RIBOLLA GIALLA"),
        "žametovka": lookup_grape("ZAMETOVKA"),
        "refosco": lookup_grape("REFOSCO DAL PEDUNCOLO ROSSO"),
        "solaris": lookup_grape("SOLARIS"),
        "johanniter": lookup_grape("JOHANNITER"),
    }

    def find_grape_id(anderson_name: str) -> str | None:
        lower = anderson_name.lower().strip()
        # Check overrides first
        if lower in GRAPE_OVERRIDES:
            return GRAPE_OVERRIDES[lower]
        # Direct lookup
        if lower in grape_lookup:
            return grape_lookup[lower]
        # Try without accents
        no_accent = normalize(lower)
        if no_accent in grape_lookup:
            return grape_lookup[no_accent]
        return None

    # 2. Parse Anderson data
    print("\nParsing Anderson & Aryal regional data...")
    sheets = read_xlsx_sheets(args.regional)

    # Aggregate by our region IDs
    region_agg: dict[str, dict[str, dict]] = {}  # region_id -> { grape_id -> {hectares, name} }
    country_agg: dict[str, dict[str, dict]] = {}  # country_id -> { grape_id -> {hectares, name} }
    unmatched_grapes: set[str] = set()
    unmatched_regions: set[str] = set()

    for sheet_name, rows in sheets.items():
        iso = COUNTRY_MAP.get(sheet_name)
        if not iso:
            continue

        country = country_by_iso.get(iso)
        if not country:
            continue

        cid = country["id"]
        if cid not in country_agg:
            country_agg[cid] = {}

        for row in rows:
            grape = row.get("prime")
            area = 0.0
            try:
                area = float(row.get("area", 0) or 0)
            except (ValueError, TypeError):
                pass
            region_name = row.get("region", "Unknown") or "Unknown"

            if not grape or area <= 0 or grape in ("other", "other red", "other white"):
                continue

            # Find grape
            grape_id = find_grape_id(grape)
            if not grape_id:
                unmatched_grapes.add(grape)
                continue

            # Country aggregation
            if grape_id not in country_agg[cid]:
                country_agg[cid][grape_id] = {"hectares": 0, "name": grape}
            country_agg[cid][grape_id]["hectares"] += area

            # Region mapping - normalize newlines
            normalized_region = region_name.replace("\n", " ").strip()
            map_key = f"{iso}|{region_name}"
            map_key_norm = f"{iso}|{normalized_region}"
            mapped_region_name = REGION_MAP.get(map_key)
            if mapped_region_name is None:
                mapped_region_name = REGION_MAP.get(map_key_norm)

            # Check if explicitly skipped (value is None in map)
            if map_key in REGION_MAP and REGION_MAP[map_key] is None:
                continue
            if map_key_norm in REGION_MAP and REGION_MAP[map_key_norm] is None:
                continue

            if mapped_region_name is None and region_name != "Unknown":
                unmatched_regions.add(f"{sheet_name} > {region_name}")
                continue

            if mapped_region_name:
                regions_for_country = region_lookup.get(cid, {})
                region = regions_for_country.get(mapped_region_name.lower())
                if region:
                    rid = region["id"]
                    if rid not in region_agg:
                        region_agg[rid] = {}
                    if grape_id not in region_agg[rid]:
                        region_agg[rid][grape_id] = {"hectares": 0, "name": grape}
                    region_agg[rid][grape_id]["hectares"] += area
                else:
                    unmatched_regions.add(f"{sheet_name} > {region_name} -> {mapped_region_name} (not found in DB)")

    # 3. Generate inserts
    print("\nGenerating inserts...")

    # Country grapes
    country_inserts = []
    for country_id, grapes in country_agg.items():
        sorted_grapes = sorted(grapes.items(), key=lambda x: -x[1]["hectares"])
        total_ha = sum(g["hectares"] for _, g in sorted_grapes)
        if total_ha <= 0:
            continue

        filtered = [
            (gid, g) for gid, g in sorted_grapes
            if g["hectares"] >= MIN_HA_COUNTRY and (g["hectares"] / total_ha) >= MIN_SHARE_COUNTRY
        ][:MAX_GRAPES_COUNTRY]

        for grape_id, g in filtered:
            share = round((g["hectares"] / total_ha) * 1000) / 10
            country_inserts.append({
                "country_id": country_id,
                "grape_id": grape_id,
                "association_type": "typical",
                "notes": f"{round(g['hectares'])} ha ({share}% of national plantings). Source: Anderson & Aryal 2023",
            })

    # Region grapes
    region_inserts = []
    for region_id, grapes in region_agg.items():
        sorted_grapes = sorted(grapes.items(), key=lambda x: -x[1]["hectares"])
        total_ha = sum(g["hectares"] for _, g in sorted_grapes)
        if total_ha <= 0:
            continue

        filtered = [
            (gid, g) for gid, g in sorted_grapes
            if g["hectares"] >= MIN_HA_REGION and (g["hectares"] / total_ha) >= MIN_SHARE_REGION
        ][:MAX_GRAPES_REGION]

        for grape_id, g in filtered:
            share = round((g["hectares"] / total_ha) * 1000) / 10
            region_inserts.append({
                "region_id": region_id,
                "grape_id": grape_id,
                "association_type": "typical",
                "notes": f"{round(g['hectares'])} ha ({share}% of regional plantings). Source: Anderson & Aryal 2023",
            })

    print(f"\nCountry grape entries: {len(country_inserts)}")
    print(f"Region grape entries: {len(region_inserts)}")
    print(f"Unmatched grapes ({len(unmatched_grapes)}): {', '.join(sorted(unmatched_grapes))}")
    print(f"Unmatched regions ({len(unmatched_regions)}):")
    for r in sorted(unmatched_regions):
        print(f"  {r}")

    if args.dry_run:
        print("\n(DRY RUN -- not writing to DB)")
        return

    # 4. Write to DB
    print("\nClearing existing data...")

    try:
        sb.table("country_grapes").delete().neq("country_id", "00000000-0000-0000-0000-000000000000").execute()
        print("  Cleared country_grapes")
    except Exception as e:
        print(f"Error clearing country_grapes: {e}")

    try:
        sb.table("region_grapes").delete().neq("region_id", "00000000-0000-0000-0000-000000000000").execute()
        print("  Cleared region_grapes")
    except Exception as e:
        print(f"Error clearing region_grapes: {e}")

    # Insert country_grapes in batches
    print("\nInserting country_grapes...")
    for i in range(0, len(country_inserts), 500):
        batch = country_inserts[i:i + 500]
        try:
            sb.table("country_grapes").upsert(batch, on_conflict="country_id,grape_id").execute()
            print(f"  Inserted {min(i + 500, len(country_inserts))}/{len(country_inserts)}")
        except Exception as e:
            print(f"  Batch {i} error: {e}")

    # Insert region_grapes in batches
    print("\nInserting region_grapes...")
    for i in range(0, len(region_inserts), 500):
        batch = region_inserts[i:i + 500]
        try:
            sb.table("region_grapes").upsert(batch, on_conflict="region_id,grape_id").execute()
            print(f"  Inserted {min(i + 500, len(region_inserts))}/{len(region_inserts)}")
        except Exception as e:
            print(f"  Batch {i} error: {e}")

    print("\nDone!")

    # Save report
    report = {
        "source": "Anderson, Nelgen & Puga (2023)",
        "url": "https://economics.adelaide.edu.au/wine-economics/databases",
        "country_grapes": len(country_inserts),
        "region_grapes": len(region_inserts),
        "unmatched_grapes": sorted(unmatched_grapes),
        "unmatched_regions": sorted(unmatched_regions),
    }
    report_path = PROJECT_ROOT / "data" / "anderson_aryal_seed_report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Report saved to {report_path}")


if __name__ == "__main__":
    main()
